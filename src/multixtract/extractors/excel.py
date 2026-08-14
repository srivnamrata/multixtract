"""Spreadsheet extractor for ``.xlsx`` / ``.xlsm`` / ``.csv``.

Ported from the ZF CAE RAG ``parse_xlsm`` parser. Each sheet becomes one
*page* whose text is a row-oriented ``key: value | key: value`` rendering
(robust for wide / sparse engineering sheets) prefixed by a sheet header
(name, column list, row count). The shared :func:`chunk_document` then windows
that text into ~500-token chunks, so chunk sizing lives in one place instead
of being re-implemented per format.

Embedded images (``.xlsx`` / ``.xlsm`` only) are pulled from ``xl/media/`` and
best-effort mapped to their sheet via the workbook/worksheet/drawing
relationship chain, then filtered through :class:`ImageFilterPipeline`.
EMF/WMF/SVG and JPEG-XR (.wdp) reuse the shared image converters.

Requires the ``[xlsx]`` extra (openpyxl) for Excel; ``.csv`` uses only the
standard library.
"""
from __future__ import annotations

import csv as _csv
import io
import logging
import os
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from typing import Any, Dict, List, Optional, Set, Tuple

from ..filters import ImageFilterPipeline
from ._image_utils import (
    RASTER_EXTS,
    VECTOR_EXTS,
    WDP_EXTS,
    batch_convert_vectors_to_png,
    decode_wdp_to_png,
    ensure_rgb_png,
)

log = logging.getLogger("multixtract.extractors.excel")

_PKG_REL_NS    = "http://schemas.openxmlformats.org/package/2006/relationships"
_SS_MAIN_NS    = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

_MAX_HEADER_COLS    = 20      # truncate the column preview in the sheet header
_MAX_ROWS_PER_SHEET = 10_000  # cap to keep memory bounded on giant sheets
_MAX_COLS_READ      = 500     # cap columns read per row


def _trim_trailing_empty_cols(rows: List[Tuple]) -> List[Tuple]:
    """Drop trailing all-None columns so sparse sheets don't report 16 K columns.

    Excel stores the sheet's 'dimension' as the bounding rectangle of ever-touched
    cells; sheets that once had data in column XFD (16 384) report 16 384 columns
    even when 16 350 of them are now empty.  We find the rightmost column that has
    at least one non-None, non-blank value and truncate all rows there.
    Returns an empty list when every cell in every row is blank.
    """
    if not rows:
        return rows
    max_col = 0
    for row in rows:
        for j in range(len(row) - 1, -1, -1):
            if row[j] is not None and str(row[j]).strip() != "":
                max_col = max(max_col, j + 1)
                break
    if max_col == 0:
        return []
    return [row[:max_col] for row in rows]


def _hidden_col_indices(ws) -> Set[int]:
    """Return the 0-based column indices that are hidden in *ws*.

    Requires the workbook to be opened with ``read_only=False``; returns an
    empty set when ``column_dimensions`` is not available.
    """
    try:
        col_dims = ws.column_dimensions
    except AttributeError:
        return set()

    try:
        from openpyxl.utils import column_index_from_string
    except ImportError:
        return set()

    hidden: Set[int] = set()
    for col_letter, dim in col_dims.items():
        if getattr(dim, "hidden", False):
            try:
                hidden.add(column_index_from_string(col_letter) - 1)  # 0-based
            except Exception:
                pass
    return hidden


def _extract_hyperlinks(ws) -> List[str]:
    """Collect hyperlink URLs from a worksheet.

    Uses the ``ws.hyperlinks`` collection (openpyxl ≥ 2.6) as the primary
    source.  Falls back to per-cell ``.hyperlink`` attribute iteration only
    when the collection is empty or unavailable, and caps that scan at
    ``_MAX_ROWS_PER_SHEET`` rows to bound cost.  Returns a deduplicated list.
    """
    seen: Set[str] = set()
    urls: List[str] = []

    def _add(url: str) -> None:
        url = (url or "").strip()
        if url and url not in seen:
            seen.add(url)
            urls.append(url)

    # Probe accessibility first so a mid-iteration error (rare) doesn't trigger
    # the expensive fallback unnecessarily.
    try:
        hyperlinks_col = ws.hyperlinks  # raises AttributeError on old openpyxl
    except Exception:
        hyperlinks_col = None

    if hyperlinks_col is not None:
        try:
            for hl in hyperlinks_col:
                _add(getattr(hl, "target", "") or "")
        except Exception:
            pass
    else:
        # Per-cell fallback for very old openpyxl or non-standard worksheet types.
        try:
            for i, row in enumerate(ws.iter_rows()):
                if i >= _MAX_ROWS_PER_SHEET:
                    break
                for cell in row:
                    hl = getattr(cell, "hyperlink", None)
                    if hl is not None:
                        _add(getattr(hl, "target", "") or str(hl))
        except Exception:
            pass

    return urls


def _named_ranges_text(wb) -> str:
    """Build a plain-text block of all workbook-level named ranges.

    Named ranges (e.g. ``INPUT_RANGE``, ``OUTPUT_TABLE``) are defined in
    ``wb.defined_names`` and are invisible to normal cell iteration.  Including
    them makes the names and their cell references searchable in the index.
    Returns an empty string when none are defined or the attribute is absent.
    """
    try:
        defined = wb.defined_names
    except AttributeError:
        return ""

    lines: List[str] = []
    try:
        items = list(defined.items())
    except Exception:
        return ""

    for name, defn in items:
        try:
            # attr_text is the cell reference string (e.g. "Sheet1!$A$1:$B$10").
            # Fall back to the value attribute, then skip if neither is a plain string.
            dest = getattr(defn, "attr_text", None) or getattr(defn, "value", None)
            if not dest or not isinstance(dest, str):
                continue
            lines.append(f"{name}: {dest}")
        except Exception:
            pass

    if not lines:
        return ""
    return "Named Ranges:\n" + "\n".join(lines)


def _row_to_kv(headers: List[str], row: Tuple) -> str:
    """Render one row as ``col: val | col: val`` for non-empty cells only."""
    pairs = []
    for h, v in zip(headers, row):
        if v is not None and str(v).strip() != "":
            pairs.append(f"{h}: {v}")
    return " | ".join(pairs)


def _is_metadata_row(row) -> bool:
    """Return True if this row looks like a key-value metadata row rather than a column header.

    Engineering spreadsheets frequently have metadata rows above the real header
    (e.g. 'Report Date: 2024-01-01', 'Filter: Active'). The reliable signal is
    that at least one non-empty cell ends with ':' — a label cell pattern.
    Real column-header rows like ('Part', 'Quantity', 'Price') never end with ':'.
    """
    non_empty = [s for c in row if c is not None and (s := str(c).strip()) != ""]
    if not non_empty:
        return False
    return any(c.endswith(":") for c in non_empty)


def _sheet_to_text(sheet_name: str, rows: List[Tuple], truncated: bool = False) -> str:
    """Build header + row-oriented text for one sheet.

    Rows are separated by a blank line so the text chunker treats each as its
    own unit.  When ``truncated`` is True a note is appended to the row count
    line to indicate that the sheet exceeded ``_MAX_ROWS_PER_SHEET``.
    """
    if not rows:
        return ""

    headers: Optional[List[str]] = None
    data_start = 0
    for i, row in enumerate(rows):
        if any(c is not None and str(c).strip() != "" for c in row):
            if _is_metadata_row(row):
                continue  # skip label:value metadata rows above the real header
            headers = [str(c) if c not in (None, "") else f"col_{j}" for j, c in enumerate(row)]
            data_start = i + 1
            break
    if not headers:
        return ""

    data_rows = rows[data_start:]
    col_preview = ", ".join(headers[:_MAX_HEADER_COLS])
    if len(headers) > _MAX_HEADER_COLS:
        col_preview += f", ... ({len(headers)} columns total)"
    row_label = f"Total rows: {len(data_rows)}"
    if truncated:
        row_label += f" (truncated to {_MAX_ROWS_PER_SHEET})"
    header_block = (
        f"Sheet: {sheet_name}\n"
        f"Columns: {col_preview}\n"
        f"{row_label}"
    )

    row_texts = [t for t in (_row_to_kv(headers, r) for r in data_rows) if t]
    if not row_texts:
        return header_block
    return header_block + "\n\n" + "\n\n".join(row_texts)


def _build_sheet_media_map(zf: zipfile.ZipFile) -> Dict[str, List[str]]:
    """Best-effort map of sheet name -> [xl/media/* paths] via the
    workbook -> worksheet -> drawing relationship chain."""
    names = set(zf.namelist())
    sheet_media: Dict[str, List[str]] = defaultdict(list)

    def _read_rels(rels_path: str) -> Dict[str, str]:
        out: Dict[str, str] = {}
        if rels_path not in names:
            return out
        try:
            root = ET.fromstring(zf.read(rels_path).decode("utf-8"))
            for rel in root.findall(f"{{{_PKG_REL_NS}}}Relationship"):
                out[rel.get("Id", "")] = rel.get("Target", "")
        except Exception as exc:
            log.debug("rels parse failed for %s: %s", rels_path, exc)
        return out

    try:
        wb_root = ET.fromstring(zf.read("xl/workbook.xml").decode("utf-8"))
    except Exception as exc:
        log.warning("could not parse workbook.xml for image mapping: %s", exc)
        return sheet_media
    wb_rels = _read_rels("xl/_rels/workbook.xml.rels")

    for sheet_el in wb_root.iter(f"{{{_SS_MAIN_NS}}}sheet"):
        sheet_name = sheet_el.get("name", "")
        rid        = sheet_el.get(f"{{{_OFFICE_REL_NS}}}id", "")
        target     = wb_rels.get(rid, "")
        if not sheet_name or not target:
            continue
        sheet_part = "xl/" + target.lstrip("/") if not target.startswith("xl/") else target
        base       = os.path.basename(sheet_part)
        sheet_rels = _read_rels(f"xl/worksheets/_rels/{base}.rels")

        for rel_target in sheet_rels.values():
            if "drawing" not in rel_target:
                continue
            drawing_name = os.path.basename(rel_target)
            drawing_rels = _read_rels(f"xl/drawings/_rels/{drawing_name}.rels")
            for media_target in drawing_rels.values():
                if "media/" in media_target:
                    sheet_media[sheet_name].append(
                        f"xl/media/{os.path.basename(media_target)}"
                    )
    return sheet_media


class ExcelExtractor:
    """DocumentExtractor for ``.xlsx`` / ``.xlsm`` / ``.csv``."""

    extensions: Tuple[str, ...] = (".xlsx", ".xlsm", ".csv")

    def __init__(self, vector_timeout: int = 120) -> None:
        self.vector_timeout = vector_timeout

    def extract(
        self,
        path: str,
        image_filter: Optional[ImageFilterPipeline] = None,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        base_name = os.path.splitext(os.path.basename(path))[0]
        empty: Dict[str, Any] = {"_base_name": base_name, "metadata": {}, "pgs": []}
        try:
            if path.lower().endswith(".csv"):
                return self._extract_csv(path, base_name)
            if image_filter is None:
                image_filter = ImageFilterPipeline()
            image_filter.reset()
            return self._extract_xlsx(path, base_name, image_filter)
        except ImportError:
            raise
        except Exception:
            log.warning("ExcelExtractor failed for %s", path, exc_info=True)
            return empty, []

    # ------------------------------------------------------------------ CSV
    def _extract_csv(self, path: str, base_name: str) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:  # noqa: E501
        with open(path, "r", encoding="utf-8", errors="replace", newline="") as f:
            sample = f.read(8192)
            f.seek(0)
            try:
                dialect = (
                    _csv.Sniffer().sniff(sample, delimiters=",;\t|")
                    if sample else _csv.excel
                )
            except _csv.Error:
                dialect = _csv.excel
            rows: List[Tuple] = []
            truncated = False
            for r in _csv.reader(f, dialect):
                if len(rows) >= _MAX_ROWS_PER_SHEET:
                    truncated = True
                    break
                rows.append(tuple(r))

        txt = _sheet_to_text(base_name, rows, truncated=truncated)
        document = {
            "metadata": {"sheet_count": 1, "row_count": max(0, len(rows) - 1)},
            "_base_name": base_name,
            "pgs": [{"pg_num": 1, "kind": "sheet", "title": base_name,
                     "txt": txt, "tables": [], "hyperlinks": [], "imgs": []}],
        }
        return document, []

    # ----------------------------------------------------------------- XLSX
    def _extract_xlsx(
        self,
        path: str,
        base_name: str,
        image_filter: ImageFilterPipeline,
    ) -> Tuple[Dict[str, Any], List[Dict[str, Any]]]:
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError(
                "Excel support requires openpyxl: pip install 'multixtract[xlsx]'"
            ) from e
        from PIL import Image

        # read_only=False is required to access sheet_state and column_dimensions.
        wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        try:
            all_names = list(wb.sheetnames)

            # Open each worksheet once to read sheet_state; keep (name, ws) pairs
            # for visible sheets so the processing loop reuses the same objects.
            visible_sheets: List[Tuple[str, Any]] = []
            for n in all_names:
                ws = wb[n]
                if getattr(ws, "sheet_state", "visible") == "visible":
                    visible_sheets.append((n, ws))

            visible_names = [n for n, _ in visible_sheets]

            # Map sheet name → pg_num using only visible sheets so that the
            # pg_num here matches the pg_num assigned in document["pgs"] below.
            sheet_index: Dict[str, int] = {
                name: i + 1 for i, name in enumerate(visible_names)
            }

            named_ranges_txt = _named_ranges_text(wb)

            document: Dict[str, Any] = {
                "metadata": {
                    "sheet_count":        len(visible_names),
                    "sheet_names":        visible_names,
                    "hidden_sheet_count": len(all_names) - len(visible_names),
                    "has_named_ranges":   bool(named_ranges_txt),
                },
                "_base_name": base_name,
                "pgs": [],
            }
            for pg_num, (name, ws) in enumerate(visible_sheets, start=1):

                # Determine hidden columns (0-based indices) before iterating rows.
                hidden_cols = _hidden_col_indices(ws)

                # Apply row and column caps; drop hidden columns in the same pass.
                raw_rows: List[Tuple] = []
                truncated = False
                for i, row in enumerate(ws.iter_rows(values_only=True)):
                    if i >= _MAX_ROWS_PER_SHEET:
                        truncated = True
                        break
                    capped: Tuple = row[:_MAX_COLS_READ]
                    if hidden_cols:
                        capped = tuple(v for j, v in enumerate(capped) if j not in hidden_cols)
                    raw_rows.append(capped)

                rows = _trim_trailing_empty_cols(raw_rows)
                hyperlinks = _extract_hyperlinks(ws)
                sheet_txt = _sheet_to_text(name, rows, truncated=truncated)

                # Append named ranges to the first sheet so they are indexed
                # once per workbook without creating an artificial extra page.
                if pg_num == 1 and named_ranges_txt:
                    sheet_txt = (
                        f"{sheet_txt}\n\n{named_ranges_txt}" if sheet_txt else named_ranges_txt
                    )

                document["pgs"].append({
                    "pg_num":     pg_num,
                    "kind":       "sheet",
                    "title":      name,
                    "txt":        sheet_txt,
                    "tables":     [],
                    "hyperlinks": hyperlinks,
                    "imgs":       [],
                })
        finally:
            wb.close()

        prepared_images: List[Dict[str, Any]] = []
        try:
            zf = zipfile.ZipFile(path, "r")
        except Exception as exc:
            log.warning("could not open %s as ZIP for image extraction: %s", base_name, exc)
            return document, prepared_images

        try:
            sheet_media_map = _build_sheet_media_map(zf)
            vector_items, wdp_items, seen = [], [], set()
            for media_paths in sheet_media_map.values():
                for media_path in media_paths:
                    if media_path in seen:
                        continue
                    ext = os.path.splitext(media_path)[1].lower()
                    try:
                        raw = zf.read(media_path)
                    except KeyError:
                        continue
                    if ext in VECTOR_EXTS:
                        vector_items.append((media_path, raw))
                        seen.add(media_path)
                    elif ext in WDP_EXTS:
                        wdp_items.append((media_path, raw))
                        seen.add(media_path)
            converted: Dict[str, bytes] = batch_convert_vectors_to_png(
                vector_items, self.vector_timeout
            )
            vector_items.clear()
            converted.update(decode_wdp_to_png(wdp_items))
            wdp_items.clear()

            per_sheet_idx: Dict[int, int] = defaultdict(int)
            processed_media: Set[str] = set()
            for sheet_name, media_paths in sheet_media_map.items():
                pg_num = sheet_index.get(sheet_name)
                if pg_num is None:
                    # Hidden sheet — no page was created for it; skip its images.
                    continue
                for media_path in media_paths:
                    # Deduplicate across all sheets (covers converted vectors
                    # and any raster referenced by multiple sheets).
                    if media_path in processed_media:
                        continue
                    ext = os.path.splitext(media_path)[1].lower()
                    if media_path in converted:
                        image_bytes, ext_out = converted.pop(media_path), "png"
                    elif ext in RASTER_EXTS:
                        try:
                            image_bytes = zf.read(media_path)
                        except KeyError:
                            continue
                        ext_out = ext.lstrip(".")
                        ext_out = {"tif": "tiff", "jpg": "jpeg"}.get(ext_out, ext_out)
                        if ext_out == "png" and not image_bytes[:4].startswith(b"\x89PNG"):
                            fixed = ensure_rgb_png(image_bytes)
                            if fixed is None:
                                continue
                            image_bytes = fixed
                    else:
                        continue

                    processed_media.add(media_path)

                    try:
                        with Image.open(io.BytesIO(image_bytes)) as image:
                            width, height = image.size
                    except Exception as exc:
                        log.debug("image decode failed for %s in sheet %r of %s: %s",
                                  media_path, sheet_name, base_name, exc)
                        continue

                    img_idx = per_sheet_idx[pg_num]
                    prepared = image_filter.prepare_image(
                        image_bytes=image_bytes,
                        ext=ext_out,
                        width=width,
                        height=height,
                        image_id=f"page_{pg_num}_img_{img_idx}",
                        page_number=pg_num,
                        img_idx=img_idx,
                    )
                    if prepared is not None:
                        prepared_images.append(prepared)
                        per_sheet_idx[pg_num] += 1  # only advance for kept images
        finally:
            zf.close()

        return document, prepared_images
